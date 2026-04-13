# -*- coding: utf-8 -*-

"""

资产管理API端点

提供资产的CRUD操作、导出、统计等功能

"""

from typing import List, Optional, Dict, Any

import json

import io

from datetime import datetime

from pydantic import BaseModel, Field



from fastapi import APIRouter, Depends, HTTPException, Query, Response, UploadFile, File, Form

from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session

from sqlalchemy import func, desc, asc, or_, and_



from app.core.deps import get_db, get_current_active_user

from app.models.user import User

from app.models.asset import Asset as AssetModel, AssetStatus, NetworkLocation

from app.schemas.asset import AssetCreate, AssetUpdate, Asset as AssetSchema, AssetList, AssetSingleConfirmRequest



# 导入用于导出的库

try:

    import pandas as pd

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment

    EXPORT_ENABLED = True

except ImportError:

    EXPORT_ENABLED = False

    print("Warning: pandas/openpyxl not installed, export functionality disabled")



router = APIRouter()



# =================== 基础CRUD操作 ===================



@router.get("/", response_model=AssetList)

async def list_assets(

    skip: int = Query(0, ge=0),

    limit: int = Query(100, ge=1, le=1000),

    search: Optional[str] = Query(None, description="搜索关键词"),

    asset_type: Optional[str] = Query(None, description="资产类型筛选"),

    network_location: Optional[NetworkLocation] = Query(None, description="网络位置筛选"),

    status: Optional[AssetStatus] = Query(None, description="状态筛选"),

    department: Optional[str] = Query(None, description="部门筛选"),

    sort_by: str = Query("created_at", description="排序字段"),

    sort_order: str = Query("desc", description="排序顺序 asc/desc"),

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """获取资产列表"""

    try:

        query = db.query(AssetModel)

        

        # 搜索过滤

        if search:

            search_filter = or_(

                AssetModel.name.contains(search),

                AssetModel.ip_address.contains(search),

                AssetModel.hostname.contains(search),

                AssetModel.device_model.contains(search),

                AssetModel.department.contains(search),

                AssetModel.notes.contains(search)

            )

            query = query.filter(search_filter)

        

        # 类型过滤

        if asset_type:

            query = query.filter(AssetModel.asset_type == asset_type)

        

        # 网络位置过滤

        if network_location:

            query = query.filter(AssetModel.network_location == network_location)

        

        # 状态过滤

        if status:

            query = query.filter(AssetModel.status == status)

        

        # 部门过滤

        if department:

            query = query.filter(AssetModel.department == department)

        

        # 排序

        if hasattr(AssetModel, sort_by):

            if sort_order.lower() == "asc":

                query = query.order_by(asc(getattr(AssetModel, sort_by)))

            else:

                query = query.order_by(desc(getattr(AssetModel, sort_by)))

        

        # 获取总数（用于分页）

        total_query = query

        total_count = total_query.count()

        

        # 分页

        assets = query.offset(skip).limit(limit).all()

        

        return AssetList(

            items=assets,

            total=total_count,

            page=skip // limit + 1 if limit > 0 else 1,

            per_page=limit,

            pages=(total_count + limit - 1) // limit if limit > 0 else 1

        )

        

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"获取资产列表失败: {str(e)}")



@router.get("/count")

async def count_assets(

    search: Optional[str] = Query(None),

    asset_type: Optional[str] = Query(None),

    network_location: Optional[NetworkLocation] = Query(None),

    status: Optional[AssetStatus] = Query(None),

    department: Optional[str] = Query(None),

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """获取资产总数（用于分页）"""

    try:

        query = db.query(func.count(AssetModel.id))

        

        # 应用相同的筛选条件

        if search:

            search_filter = or_(

                AssetModel.name.contains(search),

                AssetModel.ip_address.contains(search),

                AssetModel.hostname.contains(search),

                AssetModel.device_model.contains(search),

                AssetModel.department.contains(search),

                AssetModel.notes.contains(search)

            )

            query = query.filter(search_filter)

        

        if asset_type:

            query = query.filter(AssetModel.asset_type == asset_type)

        

        if network_location:

            query = query.filter(AssetModel.network_location == network_location)

        

        if status:

            query = query.filter(AssetModel.status == status)

        

        if department:

            query = query.filter(AssetModel.department == department)

        

        total = query.scalar()

        return {"total": total}

        

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"获取资产统计失败: {str(e)}")





@router.get("/statistics")

async def get_simple_asset_statistics(

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """获取简单的资产统计信息"""

    try:

        # 总数统计

        total_assets = db.query(func.count(AssetModel.id)).scalar()

        

        # 按类型统计

        server_count = db.query(func.count(AssetModel.id)).filter(AssetModel.asset_type == 'server').scalar()

        network_count = db.query(func.count(AssetModel.id)).filter(AssetModel.asset_type == 'network').scalar()

        

        # 按状态统计  

        active_count = db.query(func.count(AssetModel.id)).filter(AssetModel.status == 'active').scalar()

        

        return {

            "total": total_assets,

            "by_type": {

                "server": server_count,

                "network": network_count

            },

            "by_status": {

                "active": active_count

            }

        }

        

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"获取资产统计失败: {str(e)}")



@router.get("/{asset_id}", response_model=AssetSchema)

async def get_asset(

    asset_id: int,

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """获取单个资产详情"""

    try:

        asset = db.query(AssetModel).filter(AssetModel.id == asset_id).first()

        if not asset:

            raise HTTPException(status_code=404, detail="资产不存在")

        

        return asset

        

    except HTTPException:

        raise

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"获取资产详情失败: {str(e)}")



@router.post("/", response_model=AssetSchema)

async def create_asset(

    asset: AssetCreate,

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """创建新资产"""

    try:

        # 检查IP地址是否已存在（如果提供了IP）

        if asset.ip_address:

            existing_asset = db.query(AssetModel).filter(AssetModel.ip_address == asset.ip_address).first()

            if existing_asset:

                raise HTTPException(status_code=400, detail=f"IP地址 {asset.ip_address} 已被使用")

        

        # 创建资产

        db_asset = AssetModel(

            **asset.dict(),

            creator_id=current_user.id

        )

        

        db.add(db_asset)

        db.commit()

        db.refresh(db_asset)

        

        return db_asset

        

    except HTTPException:

        raise

    except Exception as e:

        db.rollback()

        raise HTTPException(status_code=500, detail=f"创建资产失败: {str(e)}")



@router.put("/{asset_id}", response_model=AssetSchema)

async def update_asset(

    asset_id: int,

    asset_update: AssetUpdate,

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """更新资产信息"""

    try:

        # 获取现有资产

        db_asset = db.query(AssetModel).filter(AssetModel.id == asset_id).first()

        if not db_asset:

            raise HTTPException(status_code=404, detail="资产不存在")

        

        # 检查IP地址冲突（如果更新了IP）

        if asset_update.ip_address and asset_update.ip_address != db_asset.ip_address:

            existing_asset = db.query(AssetModel).filter(

                and_(AssetModel.ip_address == asset_update.ip_address, AssetModel.id != asset_id)

            ).first()

            if existing_asset:

                raise HTTPException(status_code=400, detail=f"IP地址 {asset_update.ip_address} 已被使用")

        

        # 更新字段

        update_data = asset_update.dict(exclude_unset=True)

        for field, value in update_data.items():

            setattr(db_asset, field, value)

        

        # 更新时间戳

        from datetime import timezone, timedelta

        beijing_tz = timezone(timedelta(hours=8))

        db_asset.updated_at = datetime.now(beijing_tz)

        

        db.commit()

        db.refresh(db_asset)

        

        return db_asset

        

    except HTTPException:

        raise

    except Exception as e:

        db.rollback()

        raise HTTPException(status_code=500, detail=f"更新资产失败: {str(e)}")



@router.delete("/{asset_id}")

async def delete_asset(

    asset_id: int,

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """删除资产"""

    try:

        # 检查权限 - 只有超级管理员或创建者可以删除

        db_asset = db.query(AssetModel).filter(AssetModel.id == asset_id).first()

        if not db_asset:

            raise HTTPException(status_code=404, detail="资产不存在")

        

        if not current_user.is_superuser and db_asset.creator_id != current_user.id:

            raise HTTPException(status_code=403, detail="没有权限删除此资产")

        

        db.delete(db_asset)

        db.commit()

        

        return {"message": "资产删除成功", "asset_id": asset_id}

        

    except HTTPException:

        raise

    except Exception as e:

        db.rollback()

        raise HTTPException(status_code=500, detail=f"删除资产失败: {str(e)}")



# =================== 批量操作 ===================



class BatchDeleteRequest(BaseModel):

    asset_ids: List[int] = Field(..., description="要删除的资产ID列表")



@router.post("/batch/delete")

async def batch_delete_assets(

    request: BatchDeleteRequest,

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """批量删除资产"""

    try:

        if not current_user.is_superuser:

            raise HTTPException(status_code=403, detail="只有超级管理员可以批量删除资产")

        

        # 查找要删除的资产

        assets_to_delete = db.query(AssetModel).filter(AssetModel.id.in_(request.asset_ids)).all()

        

        if not assets_to_delete:

            raise HTTPException(status_code=404, detail="没有找到要删除的资产")

        

        deleted_count = len(assets_to_delete)

        

        # 执行删除

        db.query(AssetModel).filter(AssetModel.id.in_(request.asset_ids)).delete(synchronize_session=False)

        db.commit()

        

        return {

            "message": f"成功删除 {deleted_count} 个资产",

            "deleted_count": deleted_count,

            "deleted_ids": request.asset_ids

        }

        

    except HTTPException:

        raise

    except Exception as e:

        db.rollback()

        raise HTTPException(status_code=500, detail=f"批量删除失败: {str(e)}")



# =================== 统计功能 ===================



@router.get("/statistics/summary")

async def get_asset_statistics(

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """获取资产统计摘要"""

    try:

        # 总数统计

        total_assets = db.query(func.count(AssetModel.id)).scalar()

        

        # 按类型统计

        type_stats = db.query(

            AssetModel.asset_type, func.count(AssetModel.id).label('count')

        ).group_by(AssetModel.asset_type).all()

        

        # 按状态统计

        status_stats = db.query(

            AssetModel.status, func.count(AssetModel.id).label('count')

        ).group_by(AssetModel.status).all()

        

        # 按网络位置统计

        network_stats = db.query(

            AssetModel.network_location, func.count(AssetModel.id).label('count')

        ).group_by(AssetModel.network_location).all()

        

        # 按部门统计

        department_stats = db.query(

            AssetModel.department, func.count(AssetModel.id).label('count')

        ).group_by(AssetModel.department).all()

        

        return {

            "total_assets": total_assets,

            "by_type": [{"type": stat[0], "count": stat[1]} for stat in type_stats],

            "by_status": [{"status": stat[0], "count": stat[1]} for stat in status_stats],

            "by_network": [{"network": stat[0], "count": stat[1]} for stat in network_stats],

            "by_department": [{"department": stat[0] or "未分类", "count": stat[1]} for stat in department_stats]

        }

        

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"获取统计数据失败: {str(e)}")





# =================== 导出功能 ===================



class ExportRequest(BaseModel):

    asset_ids: List[int]

    format: str = "excel"

    fields: List[str] = ["name", "asset_type", "ip_address", "hostname", "device_model", "network_location", "department", "status"]



@router.post("/export")

async def export_assets(

    export_request: ExportRequest,

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """导出资产数据"""

    if not EXPORT_ENABLED:

        raise HTTPException(status_code=500, detail="导出功能不可用，请安装pandas和openpyxl")

    

    try:

        # 检查权限

        if not current_user.is_superuser:

            raise HTTPException(status_code=403, detail="只有管理员可以导出资产数据")

        

        if not export_request.asset_ids:

            raise HTTPException(status_code=400, detail="请选择要导出的资产")

        

        if not export_request.fields:

            raise HTTPException(status_code=400, detail="请选择要导出的字段")

        

        # 查询资产数据

        assets = db.query(AssetModel).filter(AssetModel.id.in_(export_request.asset_ids)).all()

        

        if not assets:

            raise HTTPException(status_code=404, detail="没有找到要导出的资产")

        

        # 字段映射

        field_mapping = {

            "name": "设备名称",

            "asset_type": "资产类型", 

            "device_model": "设备型号",

            "manufacturer": "制造商",

            "serial_number": "序列号",

            "ip_address": "IP地址",

            "mac_address": "MAC地址",

            "hostname": "主机名",

            "port": "端口",

            "network_location": "所处网络",

            "username": "用户名",

            "password": "密码",

            "ssh_key": "SSH密钥",

            "location": "物理位置",

            "rack_position": "机柜位置",

            "datacenter": "数据中心",

            "os_version": "操作系统",

            "cpu": "CPU信息",

            "memory": "内存信息", 

            "storage": "存储信息",

            "status": "状态",

            "department": "所属部门",

            "service_name": "服务名称",

            "application": "应用程序",

            "purpose": "用途说明",

            "notes": "备注",

            "created_at": "创建时间",

            "updated_at": "更新时间"

        }

        

        # 准备数据

        export_data = []

        for asset in assets:

            row = {}

            for field in export_request.fields:

                if hasattr(asset, field):

                    value = getattr(asset, field)

                    # 特殊处理

                    if field == 'asset_type':

                        # 资产类型转换为中文标签

                        type_labels = {

                            'server': '服务器',

                            'network': '网络设备', 

                            'storage': '存储设备',

                            'security': '安全设备',

                            'database': '数据库',

                            'application': '应用程序',

                            'other': '其他'

                        }

                        # 处理枚举值

                        actual_value = value.value if hasattr(value, 'value') else value

                        value = type_labels.get(actual_value, actual_value) if actual_value else ''

                    elif field == 'status':

                        # 状态转换为中文标签

                        status_labels = {

                            'active': '在用',

                            'inactive': '停用',

                            'maintenance': '维护中',

                            'retired': '已退役'

                        }

                        # 处理枚举值

                        actual_value = value.value if hasattr(value, 'value') else value

                        value = status_labels.get(actual_value, actual_value) if actual_value else ''

                    elif field == 'network_location':

                        # 网络位置转换为中文标签

                        location_labels = {

                            'office': '办公网',

                            'monitoring': '监控网',

                            'billing': '收费网'

                        }

                        # 处理枚举值

                        actual_value = value.value if hasattr(value, 'value') else value

                        value = location_labels.get(actual_value, actual_value) if actual_value else ''

                    elif field == 'tags':

                        # 处理JSON字段

                        try:

                            tags = json.loads(value) if value else []

                            value = ', '.join(tags) if isinstance(tags, list) else str(value)

                        except:

                            value = str(value) if value else ''

                    elif field in ['created_at', 'updated_at']:

                        if value:

                            # 如果是UTC时间，转换为北京时间显示

                            if hasattr(value, 'tzinfo') and value.tzinfo is None:

                                # 无时区信息，假设为UTC，转换为北京时间

                                from datetime import timezone, timedelta

                                beijing_tz = timezone(timedelta(hours=8))

                                value = value.replace(tzinfo=timezone.utc).astimezone(beijing_tz)

                            value = value.strftime('%Y-%m-%d %H:%M:%S')

                        else:

                            value = ''

                    else:

                        value = str(value) if value is not None else ''

                    

                    row[field_mapping.get(field, field)] = value

            export_data.append(row)

        

        # 创建DataFrame

        df = pd.DataFrame(export_data)

        

        if export_request.format.lower() == 'excel':

            return await _export_excel(df, f"资产导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

        else:

            return await _export_csv(df, f"资产导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}")

            

    except HTTPException:

        raise

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")



async def _export_excel(df: pd.DataFrame, filename: str) -> StreamingResponse:

    """导出Excel文件"""

    output = io.BytesIO()

    

    # 创建工作簿，确保支持中文

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        df.to_excel(writer, sheet_name='资产数据', index=False)

        

        # 获取工作表进行格式化

        workbook = writer.book

        worksheet = writer.sheets['资产数据']

        

        # 设置标题样式

        header_font = Font(bold=True, color='FFFFFF')

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        header_alignment = Alignment(horizontal='center', vertical='center')

        

        for col_num, column in enumerate(df.columns, 1):

            cell = worksheet.cell(row=1, column=col_num)

            cell.font = header_font

            cell.fill = header_fill

            cell.alignment = header_alignment

            

            # 自动调整列宽，改进中文字符宽度计算

            column_letter = cell.column_letter

            max_length = len(str(column))

            

            # 计算列内容的最大宽度（考虑中文字符）

            for i in range(min(len(df), 100)):

                cell_value = str(df[column].iloc[i]) if pd.notna(df[column].iloc[i]) else ''

                # 中文字符按2个字符宽度计算

                char_width = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell_value)

                max_length = max(max_length, char_width)

            

            # 设置合理的列宽范围

            adjusted_width = min(max(max_length + 2, 10), 50)

            worksheet.column_dimensions[column_letter].width = adjusted_width

    

    output.seek(0)

    

    # 使用URL编码的文件名以支持中文

    import urllib.parse

    encoded_filename = urllib.parse.quote(f"{filename}.xlsx", safe='')

    

    headers = {

        'Content-Disposition': f'attachment; filename*=UTF-8\'\'{encoded_filename}',

        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    }

    

    return StreamingResponse(io.BytesIO(output.read()), headers=headers)



async def _export_csv(df: pd.DataFrame, filename: str) -> StreamingResponse:

    """导出CSV文件"""

    output = io.StringIO()

    # 使用utf-8-sig编码以确保在Excel中正确显示中文

    df.to_csv(output, index=False, encoding='utf-8-sig')

    

    # 使用URL编码的文件名以支持中文

    import urllib.parse

    encoded_filename = urllib.parse.quote(f"{filename}.csv", safe='')

    

    headers = {

        'Content-Disposition': f'attachment; filename*=UTF-8\'\'{encoded_filename}',

        'Content-Type': 'text/csv; charset=utf-8'

    }

    

    # 转换为字节流以正确处理编码

    csv_content = output.getvalue()

    return StreamingResponse(

        io.BytesIO(csv_content.encode('utf-8-sig')), 

        headers=headers, 

        media_type='text/csv'

    )



# =================== 文件提取功能 ===================



@router.post("/test-file-upload")

async def test_file_upload(file: UploadFile = File(...)):

    """测试文件上传 - 最简化版本"""

    try:

        content = await file.read()

        return {

            "success": True,

            "filename": file.filename,

            "size": len(content),

            "content_type": file.content_type

        }

    except Exception as e:

        return {

            "success": False,

            "error": str(e)

        }



@router.post("/test-with-db")

async def test_with_db(

    file: UploadFile = File(...),

    db: Session = Depends(get_db)

):

    """测试数据库依赖"""

    try:

        content = await file.read()

        return {

            "success": True,

            "filename": file.filename,

            "size": len(content),

            "db_connected": db is not None

        }

    except Exception as e:

        return {

            "success": False,

            "error": str(e)

        }



@router.post("/test-with-auth")

async def test_with_auth(

    file: UploadFile = File(...),

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """测试用户认证依赖"""

    try:

        content = await file.read()

        return {

            "success": True,

            "filename": file.filename,

            "size": len(content),

            "user": current_user.username

        }

    except Exception as e:

        return {

            "success": False,

            "error": str(e)

        }



@router.post("/test-extractor")

async def test_extractor(

    file: UploadFile = File(...),

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """测试提取器功能"""

    import traceback

    try:

        # 步骤1: 读取文件

        content = await file.read()

        file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else 'txt'

        

        # 步骤2: 导入提取器

        from app.services.enhanced_asset_extractor import EnhancedAssetExtractor

        

        # 步骤3: 创建提取器

        extractor = EnhancedAssetExtractor()

        

        # 步骤4: 执行提取

        extracted_assets = extractor.extract_from_file(file.filename, content, file_extension)

        

        return {

            "success": True,

            "filename": file.filename,

            "size": len(content),

            "user": current_user.username,

            "extracted_count": len(extracted_assets),

            "sample_asset": extracted_assets[0] if extracted_assets else None

        }

        

    except Exception as e:

        return {

            "success": False,

            "error": str(e),

            "traceback": traceback.format_exc()

        }



@router.post("/file-extract")

async def extract_assets_from_file(

    file: UploadFile = File(...),

    auto_merge: bool = Form(True),

    merge_threshold: int = Form(80),

    use_ai: bool = Form(False),

    ai_provider: Optional[str] = Form(None),

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """从文件内容中提取资产信息 - 支持AI提取和传统规则提取"""

    import asyncio

    from concurrent.futures import ThreadPoolExecutor

    from app.services.ai.extractors.asset_extractor import AIAssetExtractor

    

    def sync_extract_traditional(filename: str, content: bytes, extension: str):

        """使用传统规则提取"""
        from app.services.enhanced_asset_extractor import EnhancedAssetExtractor
        extractor = EnhancedAssetExtractor()
        return extractor.extract_from_file(filename, content, extension)

    
    async def sync_extract_ai(filename: str, content: bytes, extension: str, provider: Optional[str]):
        """使用AI提取"""
        extractor = AIAssetExtractor()
        return await extractor.extract_from_file(filename, content, extension, provider=provider, use_fallback=True)
    
    def run_ai_extractor(filename: str, content: bytes, extension: str, provider: Optional[str]):
        """在同步环境中运行AI提取器"""
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            return loop.run_until_complete(sync_extract_ai(filename, content, extension, provider))
        finally:
            loop.close()

    try:

        # 读取文件内容

        file_content = await file.read()

        file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else 'txt'

        

        # 根据use_ai选择提取器

        if use_ai:
            with ThreadPoolExecutor() as executor:
                extracted_assets = await asyncio.get_event_loop().run_in_executor(
                    executor, run_ai_extractor, file.filename, file_content, file_extension, ai_provider
                )
            method = f"ai({ai_provider or 'default'})"
        else:
            with ThreadPoolExecutor() as executor:
                extracted_assets = await asyncio.get_event_loop().run_in_executor(
                    executor, sync_extract_traditional, file.filename, file_content, file_extension
                )
            method = "traditional"

        message = f"从文件提取到 {len(extracted_assets)} 个资产"

        

        if not extracted_assets:

            return {

                "success": False,

                "message": "未能从文件中提取到资产信息",

                "assets": [],

                "method": method

            }

        

        # 检查IP地址冲突和名称冲突

        for asset_data in extracted_assets:

            asset_data['conflicts'] = []

            asset_data['is_duplicate'] = False
            asset_data['existing_id'] = None
            asset_data['conflict_type'] = None
            asset_data['existing_asset'] = None

            if asset_data.get('ip_address'):

                try:

                    existing_asset = db.query(AssetModel).filter(AssetModel.ip_address == asset_data.get('ip_address')).first()

                    if existing_asset:

                        asset_data['conflicts'].append(f"IP地址 {asset_data.get('ip_address')} 已被设备 '{existing_asset.name}' 使用")

                    # 检查设备名称冲突（用于逐条确认）

                    existing_asset_by_name = db.query(AssetModel).filter(AssetModel.name == asset_data.get('name')).first()

                    if existing_asset_by_name:

                        asset_data['is_duplicate'] = True
                        asset_data['existing_id'] = existing_asset_by_name.id
                        asset_data['conflict_type'] = 'name_conflict'
                        # 返回现有资产的完整信息供前端编辑参考
                        asset_data['existing_asset'] = {
                            "id": existing_asset_by_name.id,
                            "name": existing_asset_by_name.name,
                            "asset_type": existing_asset_by_name.asset_type.value if hasattr(existing_asset_by_name.asset_type, 'value') else existing_asset_by_name.asset_type,
                            "device_model": existing_asset_by_name.device_model,
                            "ip_address": existing_asset_by_name.ip_address,
                            "hostname": existing_asset_by_name.hostname,
                            "username": existing_asset_by_name.username,
                            "password": existing_asset_by_name.password,
                            "network_location": existing_asset_by_name.network_location.value if hasattr(existing_asset_by_name.network_location, 'value') else existing_asset_by_name.network_location,
                            "department": existing_asset_by_name.department,
                            "status": existing_asset_by_name.status.value if hasattr(existing_asset_by_name.status, 'value') else existing_asset_by_name.status,
                            "notes": existing_asset_by_name.notes,
                            "created_at": existing_asset_by_name.created_at.isoformat() if existing_asset_by_name.created_at else None
                        }

                except Exception:

                    pass  # 忽略数据库查询错误，继续处理

        

        return {

            "success": True,

            "extracted_count": len(extracted_assets),

            "message": message,

            "assets": extracted_assets,

            "method": method

        }

        

    except Exception as e:

        return {

            "success": False,

            "error": str(e),

            "message": "文件提取失败",

            "method": method

        }



@router.post("/file-extract/single-confirm")
async def confirm_single_asset(
    request: AssetSingleConfirmRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """确认并保存单个资产到数据库（支持逐条确认）"""
    asset_data = request.asset_data
    is_duplicate = request.is_duplicate
    existing_id = request.existing_id
    
    try:

        if is_duplicate and existing_id:

            # 更新现有资产

            db_asset = db.query(AssetModel).filter(AssetModel.id == existing_id).first()

            if not db_asset:

                raise HTTPException(status_code=404, detail="资产不存在")

            

            # 更新非空字段

            for field, value in asset_data.items():

                if value and hasattr(db_asset, field):

                    setattr(db_asset, field, value)

            

            from datetime import timezone, timedelta

            beijing_tz = timezone(timedelta(hours=8))

            from datetime import datetime

            db_asset.updated_at = datetime.now(beijing_tz)

            

            action = "updated"

        else:

            # 创建新资产

            db_asset = AssetModel(

                name=asset_data.get('name', '未知设备'),

                asset_type=asset_data.get('asset_type', 'server'),

                device_model=asset_data.get('device_model'),

                ip_address=asset_data.get('ip_address'),

                hostname=asset_data.get('hostname'),

                username=asset_data.get('username'),

                password=asset_data.get('password'),

                network_location=asset_data.get('network_location', 'office'),

                department=asset_data.get('department'),

                status=asset_data.get('status', 'active'),

                notes=asset_data.get('notes'),

                confidence_score=asset_data.get('confidence_score', 100),

                creator_id=current_user.id

            )

            

            db.add(db_asset)

            db.flush()

            action = "created"

        

        db.commit()

        db.refresh(db_asset)

        

        return {

            "success": True,

            "message": f"资产{'保存成功' if action == 'created' else '更新成功'}",

            "action": action,

            "asset": {

                "id": db_asset.id,

                "name": db_asset.name,

                "ip_address": db_asset.ip_address,

                "asset_type": db_asset.asset_type,

                "status": db_asset.status,

                "created_at": db_asset.created_at.isoformat() if db_asset.created_at else None

            }

        }

        

    except Exception as e:

        db.rollback()

        raise HTTPException(status_code=500, detail=f"保存资产失败: {str(e)}")



@router.get("/ai/providers")

async def get_ai_providers(

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """获取所有可用的AI提供商列表"""

    try:

        from app.services.ai.ai_config import AIProvider, ai_config

        

        providers_info = []

        for provider in AIProvider:

            # 检查提供商是否已配置

            is_configured = False

            if provider == AIProvider.OPENAI:

                is_configured = bool(getattr(ai_config, 'OPENAI_API_KEY', ''))

            elif provider == AIProvider.ANTHROPIC:

                is_configured = bool(getattr(ai_config, 'ANTHROPIC_API_KEY', ''))

            elif provider == AIProvider.ALIBABA:

                is_configured = bool(getattr(ai_config, 'ALIBABA_API_KEY', ''))

            elif provider == AIProvider.ZHIPU:

                is_configured = bool(getattr(ai_config, 'ZHIPU_API_KEY', ''))

            elif provider == AIProvider.MINIMAX:

                is_configured = bool(getattr(ai_config, 'MINIMAX_API_KEY', ''))

            

            provider_info = {

                "name": provider.value if hasattr(provider, 'value') else str(provider),

                "display_name": {

                    AIProvider.OPENAI: "OpenAI",

                    AIProvider.ANTHROPIC: "Anthropic (Claude)",

                    AIProvider.ALIBABA: "阿里云通义",

                    AIProvider.ZHIPU: "智谱AI",

                    AIProvider.MINIMAX: "MiniMax",

                    AIProvider.MOCK: "模拟测试"

                }.get(provider, str(provider)),

                "is_available": is_configured,

                "description": {

                    AIProvider.OPENAI: "OpenAI GPT系列模型",

                    AIProvider.ANTHROPIC: "Anthropic Claude系列模型",

                    AIProvider.ALIBABA: "阿里云通义千问系列模型",

                    AIProvider.ZHIPU: "智谱AI GLM系列模型",

                    AIProvider.MINIMAX: "MiniMax ABAB系列模型",

                    AIProvider.MOCK: "用于测试的模拟提供商"

                }.get(provider, "")

            }

            

            providers_info.append(provider_info)

        

        # 添加自定义提供商选项

        providers_info.append({

            "name": "custom",

            "display_name": "自定义 (Custom)",

            "is_available": True,

            "description": "自定义AI服务提供商"

        })

        

        return {

            "success": True,

            "providers": providers_info,

            "default_provider": getattr(ai_config, 'DEFAULT_PROVIDER', 'openai').value if hasattr(getattr(ai_config, 'DEFAULT_PROVIDER', 'openai'), 'value') else str(getattr(ai_config, 'DEFAULT_PROVIDER', 'openai'))

        }

        

    except Exception as e:

        import traceback

        print(f"获取AI提供商列表失败: {str(e)}")

        print(traceback.format_exc())

        return {

            "success": False,

            "error": str(e)

        }



@router.get("/ai/stats")

async def get_ai_stats(

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """获取AI使用统计"""

    try:

        from app.services.ai.ai_service import ai_service



        stats = ai_service.get_cost_stats()



        return {

            "success": True,

            "stats": stats

        }



    except Exception as e:

        return {

            "success": False,

            "error": str(e)

        }



@router.post("/ai/test")

async def test_ai_connection(

    request: Dict[str, Any],

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """测试AI连接"""

    try:

        provider = request.get("provider", "openai")

        config_data = request.get("config", {})

        

        # 导入AI服务

        from app.services.ai.ai_config import ai_config, AIProvider

        from app.services.ai.ai_service import ai_service

        import httpx

        

        # 优先使用请求中提供的配置，否则使用默认配置

        if provider == "openai":

            api_key = config_data.get('api_key', getattr(ai_config, 'OPENAI_API_KEY', ''))

            api_url = config_data.get('api_url', getattr(ai_config, 'OPENAI_API_URL', 'https://api.openai.com/v1/chat/completions'))

            model = config_data.get('model', getattr(ai_config, 'OPENAI_MODEL', 'gpt-4o-mini'))

            

            if not api_key:

                return {

                    "success": False,

                    "message": "OpenAI API Key未配置"

                }

            

            # 测试连接

            headers = {

                "Authorization": f"Bearer {api_key}",

                "Content-Type": "application/json"

            }

            

            test_data = {

                "model": model,

                "messages": [{"role": "user", "content": "test"}],

                "max_tokens": 5

            }

            

            async with httpx.AsyncClient(timeout=10.0) as client:

                response = await client.post(api_url, json=test_data, headers=headers)

                

                if response.status_code == 200:

                    return {

                        "success": True,

                        "message": f"OpenAI连接测试成功 (模型: {model})"

                    }

                else:

                    return {

                        "success": False,

                        "message": f"OpenAI连接测试失败: HTTP {response.status_code} - {response.text}"

                    }

                    

        elif provider == "anthropic":

            api_key = config_data.get('api_key', getattr(ai_config, 'ANTHROPIC_API_KEY', ''))

            api_url = config_data.get('api_url', getattr(ai_config, 'ANTHROPIC_API_URL', 'https://api.anthropic.com/v1/messages'))

            model = config_data.get('model', getattr(ai_config, 'ANTHROPIC_MODEL', 'claude-3-haiku-20240307'))

            

            if not api_key:

                return {

                    "success": False,

                    "message": "Anthropic API Key未配置"

                }

            

            # 测试连接

            headers = {

                "x-api-key": api_key,

                "Content-Type": "application/json",

                "anthropic-version": "2023-06-01"

            }

            

            test_data = {

                "model": model,

                "max_tokens": 5,

                "messages": [{"role": "user", "content": "test"}]

            }

            

            async with httpx.AsyncClient(timeout=10.0) as client:

                response = await client.post(api_url, json=test_data, headers=headers)

                

                if response.status_code == 200:

                    return {

                        "success": True,

                        "message": f"Anthropic连接测试成功 (模型: {model})"

                    }

                else:

                    return {

                        "success": False,

                        "message": f"Anthropic连接测试失败: HTTP {response.status_code} - {response.text}"

                    }

                    

        elif provider == "alibaba":

            api_key = config_data.get('api_key', getattr(ai_config, 'ALIBABA_API_KEY', ''))

            api_url = config_data.get('api_url', getattr(ai_config, 'ALIBABA_ENDPOINT', 'https://dashscope.aliyuncs.com/compatible-mode/v1'))

            model = config_data.get('model', getattr(ai_config, 'ALIBABA_MODEL', 'qwen-plus'))

            

            if not api_key:

                return {

                    "success": False,

                    "message": "阿里云API Key未配置"

                }

            

            # 测试连接

            headers = {

                "Authorization": f"Bearer {api_key}",

                "Content-Type": "application/json"

            }

            

            test_data = {

                "model": model,

                "messages": [{"role": "user", "content": "test"}],

                "max_tokens": 5

            }

            

            async with httpx.AsyncClient(timeout=10.0) as client:

                response = await client.post(api_url, json=test_data, headers=headers)

                

                if response.status_code == 200:

                    return {

                        "success": True,

                        "message": f"阿里云通义连接测试成功 (模型: {model})"

                    }

                else:

                    return {

                        "success": False,

                        "message": f"阿里云通义连接测试失败: HTTP {response.status_code} - {response.text}"

                    }

                    

        elif provider == "zhipu":

            api_key = config_data.get('api_key', getattr(ai_config, 'ZHIPU_API_KEY', ''))

            api_url = config_data.get('api_url', getattr(ai_config, 'ZHIPU_API_URL', 'https://open.bigmodel.cn/api/paas/v4/chat/completions'))

            model = config_data.get('model', getattr(ai_config, 'ZHIPU_MODEL', 'glm-4-flash'))

            

            if not api_key:

                return {

                    "success": False,

                    "message": "智谱AI API Key未配置"

                }

            

            # 测试连接

            headers = {

                "Authorization": f"Bearer {api_key}",

                "Content-Type": "application/json"

            }

            

            test_data = {

                "model": model,

                "messages": [{"role": "user", "content": "test"}],

                "max_tokens": 5

            }

            

            async with httpx.AsyncClient(timeout=10.0) as client:

                response = await client.post(api_url, json=test_data, headers=headers)

                

                if response.status_code == 200:

                    return {

                        "success": True,

                        "message": f"智谱AI连接测试成功 (模型: {model})"

                    }

                else:

                    return {

                        "success": False,

                        "message": f"智谱AI连接测试失败: HTTP {response.status_code} - {response.text}"

                    }

                    

        elif provider == "minimax":

            api_key = config_data.get('api_key', getattr(ai_config, 'MINIMAX_API_KEY', ''))

            # 构建完整的API URL（兼容OpenAI和Anthropic端点）
            provided_url = config_data.get('api_url', '')
            
            # 检测端点类型
            if '/anthropic' in provided_url:
                # Anthropic兼容端点(Token Plan推荐)
                if '/v1/messages' in provided_url:
                    api_url = provided_url
                elif '/messages' in provided_url:
                    api_url = provided_url
                else:
                    api_url = f"{provided_url}/v1/messages"
                use_anthropic_format = True
            elif '/chat/completions' in provided_url or '/v1/text/' in provided_url:
                # OpenAI兼容端点
                api_url = provided_url
                use_anthropic_format = False
            else:
                # 默认使用Anthropic兼容端点(Token Plan推荐)
                base_url = provided_url if provided_url else "https://api.minimaxi.com/anthropic"
                api_url = f"{base_url}/v1/messages"
                use_anthropic_format = True

            model = config_data.get('model', getattr(ai_config, 'MINIMAX_MODEL', 'MiniMax-M2.7'))

            group_id = config_data.get('group_id', getattr(ai_config, 'MINIMAX_GROUP_ID', ''))



            if not api_key:

                return {

                    "success": False,

                    "message": "MiniMax API Key未配置"

                }



            # 根据端点类型构建请求

            if use_anthropic_format:
                # Anthropic格式
                headers = {
                    "x-api-key": api_key,
                    "Content-Type": "application/json",
                    "anthropic-version": "2023-06-01"
                }
                test_data = {
                    "model": model,
                    "messages": [{"role": "user", "content": "test"}],
                    "max_tokens": 5
                }
            else:
                # OpenAI格式
                headers = {
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json"
                }
                test_data = {
                    "model": model,
                    "messages": [{"role": "user", "content": "test"}],
                    "max_tokens": 5
                }
                if group_id:
                    test_data["group_id"] = group_id



            async with httpx.AsyncClient(timeout=10.0) as client:

                response = await client.post(api_url, json=test_data, headers=headers)



                if response.status_code == 200:

                    result = response.json()



                    # 检查响应中的错误

                    if result.get("base_resp"):

                        base_resp = result["base_resp"]
                        status_code = base_resp.get('status_code')
                        status_msg = base_resp.get('status_msg', '')



                        # 2013: 模型不存在

                        if status_code == 2013:

                            return {

                                "success": False,

                                "message": f"MiniMax模型 {model} 不存在。正确模型名称为: MiniMax-M2.7 或 MiniMax-M2.7-highspeed（极速版需特定订阅）"

                            }



            # 测试连接 - 使用OpenAI兼容格式（根据官方文档）

            headers = {

                "Authorization": f"Bearer {api_key}",

                "Content-Type": "application/json"

            }



            test_data = {

                "model": model,

                "messages": [{"role": "user", "content": "test"}],

                "max_tokens": 5  # 使用max_tokens

            }



            # 如果有group_id，添加到请求中

            if group_id:

                test_data["group_id"] = group_id



            async with httpx.AsyncClient(timeout=10.0) as client:

                response = await client.post(api_url, json=test_data, headers=headers)



                if response.status_code == 200:

                    result = response.json()



                    # 检查响应中的错误

                    if result.get("base_resp"):

                        base_resp = result["base_resp"]

                        status_code = base_resp.get('status_code')

                        status_msg = base_resp.get('status_msg', '')
                        # 2013: 模型不存在

                        if status_code == 2013:

                            return {

                                "success": False,

                                "message": f"MiniMax模型 {model} 不存在。正确模型名称为: MiniMax-M2.7 或 MiniMax-M2.7-highspeed（极速版需特定订阅）"

                            }



                        # 2061: 计划不支持

                        elif status_code == 2061:

                            return {

                                "success": False,

                                "message": f"MiniMax模型 {model} 存在，但您的当前计划不支持此模型。请升级您的MiniMax Token Plan或使用MiniMax-M2.7。"

                            }

                        # status_code == 0 或其他成功状态

                        elif status_code == 0 or status_code is None:

                            return {

                                "success": True,

                                "message": f"MiniMax连接测试成功 (模型: {model})"

                            }



                        else:

                            return {

                                "success": False,

                                "message": f"MiniMax API错误: {status_msg}"

                            }



                    # 如果没有base_resp，检查是否有有效的响应（OpenAI格式）

                    elif result.get("choices") and result["choices"][0].get("message", {}).get("content"):

                        return {

                            "success": True,

                            "message": f"MiniMax连接测试成功 (模型: {model})"

                        }

                    else:

                        return {

                            "success": False,

                            "message": f"MiniMax响应格式异常，请检查配置"

                        }
                else:

                    return {

                        "success": False,

                        "message": f"MiniMax连接测试失败: HTTP {response.status_code} - {response.text}"

                    }

            api_key = config_data.get('api_key', '')

            api_url = config_data.get('api_url', '')

            model = config_data.get('model', '')

            

            if not api_key:

                return {

                    "success": False,

                    "message": "API Key未配置"

                }

            

            if not api_url:

                return {

                    "success": False,

                    "message": "API URL未配置"

                }

            

            if not model:

                return {

                    "success": False,

                    "message": "模型名称未配置"

                }

            

            # 测试连接

            headers = {

                "Authorization": f"Bearer {api_key}",

                "Content-Type": "application/json"

            }

            

            test_data = {

                "model": model,

                "messages": [{"role": "user", "content": "test"}],

                "max_tokens": 5

            }

            

            async with httpx.AsyncClient(timeout=10.0) as client:

                response = await client.post(api_url, json=test_data, headers=headers)

                

                if response.status_code == 200:

                    return {

                        "success": True,

                        "message": f"自定义AI连接测试成功 (模型: {model})"

                    }

                else:

                    return {

                        "success": False,

                        "message": f"自定义AI连接测试失败: HTTP {response.status_code} - {response.text}"

                    }

        else:

            return {

                "success": False,

                "message": f"不支持的AI提供商: {provider}"

            }



    except httpx.TimeoutException:

        return {

            "success": False,

            "message": "连接超时，请检查网络或API URL是否正确"

        }

    except Exception as e:

        import traceback

        return {

            "success": False,

            "message": f"AI连接测试失败: {str(e)}"

        }



# =================== 调试功能 ===================



@router.post("/debug-excel-extract")

async def debug_excel_extract(

    file: UploadFile = File(...),

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """调试Excel提取功能"""

    import traceback

    

    try:

        print("=== 调试Excel提取开始 ===")

        

        # 步骤1: 检查文件基本信息

        print(f"文件名: {file.filename}")

        print(f"Content-Type: {file.content_type}")

        print(f"文件大小限制检查...")

        

        # 步骤2: 读取文件内容

        print("读取文件内容...")

        file_content = await file.read()

        print(f"成功读取文件，大小: {len(file_content)} 字节")

        

        # 步骤3: 检查文件扩展名

        file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else 'txt'

        print(f"文件扩展名: {file_extension}")

        

        # 步骤4: 测试导入

        print("测试导入EnhancedAssetExtractor...")

        from app.services.enhanced_asset_extractor import EnhancedAssetExtractor

        print("导入成功")

        

        # 步骤5: 创建提取器

        print("创建提取器实例...")

        extractor = EnhancedAssetExtractor()

        print("提取器创建成功")

        

        # 步骤6: 测试文件内容是否为有效Excel

        print("验证Excel文件格式...")

        if file_extension in ['xlsx', 'xls']:

            # 直接测试pandas读取

            import pandas as pd

            from io import BytesIO

            

            try:

                excel_file = BytesIO(file_content)

                engine = 'openpyxl' if file_extension == 'xlsx' else 'xlrd'

                print(f"使用引擎: {engine}")

                

                # 测试读取Excel文件

                excel_data = pd.ExcelFile(excel_file, engine=engine)

                print(f"Excel文件验证成功，工作表: {excel_data.sheet_names}")

                

            except Exception as excel_error:

                print(f"Excel文件验证失败: {str(excel_error)}")

                return {

                    "success": False,

                    "step": "Excel文件验证",

                    "error": str(excel_error),

                    "message": "Excel文件格式验证失败"

                }

        

        # 步骤7: 测试提取方法

        print("开始资产提取...")

        try:

            extracted_assets = extractor.extract_from_file(file.filename, file_content, file_extension)

            print(f"提取成功: {len(extracted_assets)} 个资产")

            

            return {

                "success": True,

                "message": "调试成功",

                "extracted_count": len(extracted_assets),

                "step": "完成",

                "debug_info": {

                    "filename": file.filename,

                    "size": len(file_content),

                    "extension": file_extension,

                    "content_type": file.content_type

                }

            }

            

        except Exception as extract_error:

            print(f"提取失败: {str(extract_error)}")

            traceback.print_exc()

            return {

                "success": False,

                "step": "资产提取",

                "error": str(extract_error),

                "traceback": traceback.format_exc(),

                "message": "资产提取失败"

            }

            

    except Exception as e:

        print(f"调试过程出现异常: {str(e)}")

        traceback.print_exc()

        return {

            "success": False,

            "step": "未知",

            "error": str(e),

            "traceback": traceback.format_exc(),

            "message": "调试过程失败"

        }



# =================== 高级功能 ===================



@router.get("/search/suggestions")

async def get_search_suggestions(

    q: str = Query(..., description="搜索关键词"),

    limit: int = Query(5, ge=1, le=20),

    db: Session = Depends(get_db),

    current_user: User = Depends(get_current_active_user)

):

    """获取搜索建议"""

    try:

        suggestions = []

        

        # IP地址建议

        ip_matches = db.query(AssetModel.ip_address).filter(

            AssetModel.ip_address.contains(q)

        ).limit(limit).all()

        suggestions.extend([{"type": "ip", "value": match[0], "label": f"IP: {match[0]}"} for match in ip_matches])

        

        # 设备名称建议

        name_matches = db.query(AssetModel.name).filter(

            AssetModel.name.contains(q)

        ).limit(limit).all()

        suggestions.extend([{"type": "name", "value": match[0], "label": f"设备: {match[0]}"} for match in name_matches])

        

        # 部门建议

        dept_matches = db.query(AssetModel.department).filter(

            and_(AssetModel.department.is_not(None), AssetModel.department.contains(q))

        ).distinct().limit(limit).all()

        suggestions.extend([{"type": "department", "value": match[0], "label": f"部门: {match[0]}"} for match in dept_matches])

        

        return suggestions[:limit]

        

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"获取搜索建议失败: {str(e)}")